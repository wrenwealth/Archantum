"""Generate BTC Up/Down Polymarket Cheat Sheet Excel file."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def make_styles():
    return {
        "title": Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
        "title_fill": PatternFill(start_color="1B1B2F", end_color="1B1B2F", fill_type="solid"),
        "header": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color="2D4059", end_color="2D4059", fill_type="solid"),
        "sub_header": Font(name="Calibri", size=11, bold=True, color="1B1B2F"),
        "sub_fill": PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid"),
        "normal": Font(name="Calibri", size=11),
        "bold": Font(name="Calibri", size=11, bold=True),
        "green": Font(name="Calibri", size=11, bold=True, color="1E8449"),
        "red": Font(name="Calibri", size=11, bold=True, color="C0392B"),
        "orange": Font(name="Calibri", size=11, bold=True, color="E67E22"),
        "green_fill": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        "red_fill": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        "orange_fill": PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid"),
        "gray_fill": PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid"),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin", color="BDC3C7"),
            right=Side(style="thin", color="BDC3C7"),
            top=Side(style="thin", color="BDC3C7"),
            bottom=Side(style="thin", color="BDC3C7"),
        ),
    }


def write_title(ws, row, title, cols, s):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = s["title"]
    cell.fill = s["title_fill"]
    cell.alignment = s["center"]
    return row + 1


def write_section(ws, row, title, cols, s):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = s["sub_header"]
    cell.fill = s["sub_fill"]
    cell.alignment = s["left"]
    return row + 1


def write_header(ws, row, headers, s):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = s["header"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["border"]
    return row + 1


def write_row(ws, row, values, s, fonts=None, fills=None):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = fonts[col - 1] if fonts else s["normal"]
        if fills and fills[col - 1]:
            cell.fill = fills[col - 1]
        cell.alignment = s["center"]
        cell.border = s["border"]
    return row + 1


def build_main_sheet(wb, s):
    ws = wb.active
    ws.title = "Cheat Sheet"
    ws.sheet_properties.tabColor = "1B1B2F"

    # Column widths
    widths = [22, 14, 14, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = write_title(ws, row, "BTC UP/DOWN POLYMARKET — CHEAT SHEET", 5, s)
    row += 1

    # ── BREAK-EVEN MATH ──
    row = write_section(ws, row, "1. BREAK-EVEN MATH", 5, s)
    row = write_header(ws, row, ["Entry Price", "Break-even WR", "Win per $10", "Loss per $10", "Catatan"], s)

    be_data = [
        ("$0.95", "95.0%", "$0.53", "-$10.00", "Hampir mustahil"),
        ("$0.92", "92.0%", "$0.87", "-$10.00", "Standard near close"),
        ("$0.90", "90.0%", "$1.11", "-$10.00", "Good entry"),
        ("$0.85", "85.0%", "$1.76", "-$10.00", "Great entry"),
        ("$0.80", "80.0%", "$2.50", "-$10.00", "Rare, golden entry"),
    ]
    for vals in be_data:
        color = s["red"] if vals[1] == "95.0%" else s["orange"] if vals[1] == "92.0%" else s["green"]
        row = write_row(ws, row, vals, s, fonts=[s["bold"], color, s["green"], s["red"], s["normal"]])

    row += 1

    # ── 15M WIN RATE ──
    row = write_section(ws, row, "2A. WIN RATE — 15 MINUTE WINDOW", 5, s)
    row = write_header(ws, row, ["Entry Timing", "WR (raw)", "Profitable?", "Gap utk 92% WR", "Verdict"], s)

    m15_data = [
        ("1 min before", "82.1%", "TIDAK", "$75", ""),
        ("2 min before", "80.5%", "TIDAK", "$100", ""),
        ("3 min before", "78.6%", "TIDAK", "$100", ""),
        ("5 min before", "76.2%", "TIDAK", "$75", ""),
        ("7 min before", "73.8%", "TIDAK", "~$150", ""),
        ("10 min before", "68.4%", "TIDAK", "~$200", ""),
    ]
    for vals in m15_data:
        row = write_row(ws, row, vals, s,
                        fonts=[s["normal"], s["red"], s["red"], s["bold"], s["normal"]],
                        fills=[None, s["red_fill"], s["red_fill"], None, None])

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value="15m raw = RUGI. Harus pakai gap filter!")
    cell.font = s["red"]
    cell.fill = s["red_fill"]
    cell.alignment = s["center"]
    row += 2

    # ── 5M WIN RATE ──
    row = write_section(ws, row, "2B. WIN RATE — 5 MINUTE WINDOW", 5, s)
    row = write_header(ws, row, ["Entry Timing", "WR (raw)", "Profitable?", "Gap utk 92% WR", "Verdict"], s)
    row = write_row(ws, row, ("1 min before", "84.9%", "TIDAK", "$30", "Butuh gap filter"),
                    s, fonts=[s["normal"], s["orange"], s["red"], s["bold"], s["normal"]],
                    fills=[None, s["orange_fill"], s["red_fill"], None, None])
    row += 1

    # ── 1H WIN RATE ──
    row = write_section(ws, row, "2C. WIN RATE — 1 HOUR WINDOW", 5, s)
    row = write_header(ws, row, ["Entry Timing", "WR (raw)", "Profitable?", "Gap utk 92% WR", "Verdict"], s)

    h1_data = [
        ("1 min before", "97.6%", "YA", "$0", "ALWAYS TRADE"),
        ("3 min before", "94.2%", "YA", "$0", "ALWAYS TRADE"),
        ("5 min before", "91.8%", "Borderline", "$50", "Pakai gap filter"),
        ("10 min before", "87.3%", "TIDAK", "$100", "Butuh gap besar"),
    ]
    for vals in h1_data:
        is_yes = vals[2] == "YA"
        is_bl = vals[2] == "Borderline"
        prof_font = s["green"] if is_yes else s["orange"] if is_bl else s["red"]
        prof_fill = s["green_fill"] if is_yes else s["orange_fill"] if is_bl else s["red_fill"]
        wr_font = s["green"] if is_yes else s["orange"] if is_bl else s["red"]
        wr_fill = s["green_fill"] if is_yes else s["orange_fill"] if is_bl else s["red_fill"]
        row = write_row(ws, row, vals, s,
                        fonts=[s["normal"], wr_font, prof_font, s["bold"], s["normal"]],
                        fills=[None, wr_fill, prof_fill, None, None])

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value="1h @ 1-3m before = PROFITABLE tanpa filter!")
    cell.font = s["green"]
    cell.fill = s["green_fill"]
    cell.alignment = s["center"]
    row += 2

    # ── GAP FILTER ──
    row = write_section(ws, row, "3. GAP FILTER — QUICK REFERENCE", 5, s)
    row = write_header(ws, row, ["Window", "Entry Timing", "Min Gap", "Win Rate", "Action"], s)

    gap_data = [
        ("15M", "1m before", "≥$75", "99.1%", "TRADE"),
        ("15M", "1m before", "≥$50", "94.8%", "TRADE"),
        ("15M", "1m before", "<$50", "~78%", "SKIP"),
        ("15M", "3m before", "≥$100", "92.3%", "TRADE"),
        ("15M", "3m before", "<$100", "~76%", "SKIP"),
        ("15M", "5m before", "≥$75", "92.0%", "TRADE (borderline)"),
        ("15M", "5m before", "<$75", "~73%", "SKIP"),
        ("1H", "1m before", "any", "97.6%", "ALWAYS TRADE"),
        ("1H", "3m before", "any", "94.2%", "ALWAYS TRADE"),
        ("1H", "5m before", "≥$50", "95.0%", "TRADE"),
        ("1H", "5m before", "<$50", "~88%", "SKIP"),
    ]
    for vals in gap_data:
        is_trade = "TRADE" in vals[4] and "SKIP" not in vals[4]
        act_font = s["green"] if is_trade else s["red"]
        act_fill = s["green_fill"] if is_trade else s["red_fill"]
        row = write_row(ws, row, vals, s,
                        fonts=[s["bold"], s["normal"], s["bold"], s["bold"], act_font],
                        fills=[None, None, None, None, act_fill])
    row += 1

    # ── PRICE SOURCE ──
    row = write_section(ws, row, "4. PRICE SOURCE — DEVIASI DARI SETTLEMENT", 5, s)
    row = write_header(ws, row, ["Source", "Vs Settlement", "Latency", "Rekomendasi", ""], s)

    ps_data = [
        ("Polymarket UI", "Baseline", "~1-5s", "Price to beat = oracle opening", ""),
        ("Binance Spot", "±$1-10", "<1s", "PAKAI INI — closest to oracle", ""),
        ("Hyperliquid", "±$5-30", "<1s", "HATI-HATI — perp, bisa deviasi", ""),
        ("TradingView", "±$5-20", "1-5s", "OK sebagai backup", ""),
        ("CoinGecko", "±$10-50", "15-60s", "Jangan pakai — terlalu delayed", ""),
    ]
    for vals in ps_data:
        rec_font = s["green"] if "PAKAI" in vals[3] else s["red"] if "HATI" in vals[3] or "Jangan" in vals[3] else s["normal"]
        row = write_row(ws, row, vals, s, fonts=[s["bold"], s["bold"], s["normal"], rec_font, s["normal"]])
    row += 1

    # ── EXPECTED P&L ──
    row = write_section(ws, row, "5. EXPECTED P&L (per 100 trades @ $10)", 5, s)
    row = write_header(ws, row, ["Strategy", "Win Rate", "Wins", "Losses", "P&L"], s)

    pnl_data = [
        ("15m raw (no filter)", "82.1%", 82, 18, "-$108.54"),
        ("15m + $50 gap", "94.8%", 95, 5, "+$32.65"),
        ("15m + $75 gap", "99.1%", 99, 1, "+$76.13"),
        ("1h raw @ 1m before", "97.6%", 98, 2, "+$65.26"),
        ("1h raw @ 3m before", "94.2%", 94, 6, "+$21.78"),
    ]
    for vals in pnl_data:
        is_profit = vals[4].startswith("+")
        pnl_font = s["green"] if is_profit else s["red"]
        pnl_fill = s["green_fill"] if is_profit else s["red_fill"]
        row = write_row(ws, row, vals, s,
                        fonts=[s["normal"], s["bold"], s["normal"], s["normal"], pnl_font],
                        fills=[None, None, None, None, pnl_fill])
    row += 1

    # ── CONFLUENCE ──
    row = write_section(ws, row, "6. CONFLUENCE — 1H TREND → 15M FILTER", 5, s)
    row = write_header(ws, row, ["Kondisi", "WR 15M (1m before)", "vs Baseline", "", ""], s)

    conf_data = [
        ("Baseline (tanpa filter)", "82.1%", "—", "", ""),
        ("1H align (searah)", "82.8%", "+0.7%", "", ""),
        ("1H counter (berlawanan)", "80.7%", "-1.4%", "", ""),
        ("1H align + $75 gap", "99.1%", "+17.0%", "", ""),
        ("1H counter + $75 gap", "~96%", "+14%", "", ""),
    ]
    for vals in conf_data:
        diff = vals[2]
        d_font = s["green"] if diff.startswith("+") else s["red"] if diff.startswith("-") else s["normal"]
        row = write_row(ws, row, vals, s, fonts=[s["normal"], s["bold"], d_font, s["normal"], s["normal"]])
    row += 1

    # ── QUARTER ANALYSIS ──
    row = write_section(ws, row, "7. QUARTER ANALYSIS (posisi 15m dalam 1 jam)", 5, s)
    row = write_header(ws, row, ["Quarter", "Menit ke-", "Align dgn 1H", "Catatan", ""], s)

    q_data = [
        ("Q1", "0-15", "82.3%", "Normal", ""),
        ("Q2", "15-30", "84.8%", "Best alignment", ""),
        ("Q3", "30-45", "31.8%", "REVERSAL zone!", ""),
        ("Q4", "45-60", "78.5%", "Normal", ""),
    ]
    for vals in q_data:
        is_rev = "REVERSAL" in vals[3]
        cat_font = s["red"] if is_rev else s["green"] if "Best" in vals[3] else s["normal"]
        cat_fill = s["red_fill"] if is_rev else s["green_fill"] if "Best" in vals[3] else None
        row = write_row(ws, row, vals, s,
                        fonts=[s["bold"], s["normal"], s["bold"], cat_font, s["normal"]],
                        fills=[None, None, None, cat_fill, None])

    return ws


def build_rules_sheet(wb, s):
    ws = wb.create_sheet("Trade Rules")
    ws.sheet_properties.tabColor = "1E8449"

    widths = [5, 45, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = write_title(ws, row, "TRADE / SKIP RULES", 3, s)
    row += 1

    # TRADE rules
    row = write_section(ws, row, "TRADE kalau:", 3, s)
    trades = [
        ("1", "1H window, 1-3m before resolve, any gap", "WR 94-97%"),
        ("2", "15M window, 1m before, gap ≥$75", "WR 99.1%"),
        ("3", "15M window, 1m before, gap ≥$50", "WR 94.8%"),
        ("4", "Entry price ≤ $0.92", ""),
    ]
    for vals in trades:
        row = write_row(ws, row, vals, s,
                        fonts=[s["green"], s["bold"], s["green"]],
                        fills=[s["green_fill"], s["green_fill"], s["green_fill"]])

    row += 1
    row = write_section(ws, row, "SKIP kalau:", 3, s)
    skips = [
        ("1", "15M window tanpa gap filter", "WR 82% = RUGI"),
        ("2", "Gap borderline (within $10 of minimum)", "Terlalu risky"),
        ("3", "Entry price > $0.95", "Break-even 95%"),
        ("4", "Sudah 3 losses berturut-turut hari ini", "Stop trading"),
        ("5", "Window di Q3 (menit 30-45 dalam jam)", "Reversal zone"),
        ("6", "Harga BTC flat/sideways (gap < $20)", "No signal"),
    ]
    for vals in skips:
        row = write_row(ws, row, vals, s,
                        fonts=[s["red"], s["bold"], s["red"]],
                        fills=[s["red_fill"], s["red_fill"], s["red_fill"]])

    row += 2

    # EXECUTION CHECKLIST
    row = write_section(ws, row, "EXECUTION CHECKLIST", 3, s)
    steps = [
        ("1", "Cek window mana yang akan resolve (15m / 1h)", ""),
        ("2", "Catat 'price to beat' dari Polymarket UI", ""),
        ("3", "Buka Binance spot, catat harga BTC sekarang", ""),
        ("4", "Hitung gap: |current - price_to_beat|", ""),
        ("5", "Cek tabel gap minimum di sheet 'Cheat Sheet'", ""),
        ("6", "Kalau gap cukup → tentukan arah (UP/DOWN)", ""),
        ("7", "Cek entry price di orderbook Polymarket", ""),
        ("8", "Kalau entry ≤ $0.92 → BUY", ""),
        ("9", "Kalau entry > $0.92 → cek break-even table", ""),
    ]
    for vals in steps:
        row = write_row(ws, row, vals, s, fonts=[s["bold"], s["normal"], s["normal"]])

    row += 2

    # COMMON MISTAKES
    row = write_section(ws, row, "COMMON MISTAKES", 3, s)
    row = write_header(ws, row, ["#", "Mistake", "Kenapa salah"], s)
    mistakes = [
        ("1", "Trade tanpa gap filter (15m)", "WR 82% = guaranteed loss"),
        ("2", "Pakai harga Hyperliquid as-is", "Bisa beda $5-30 dari oracle"),
        ("3", "Entry terlalu awal (>5m before)", "WR drop drastis"),
        ("4", "Harga entry >$0.95", "Break-even WR 95%"),
        ("5", "Ignore loss streak", "Max streak 4-6 dalam data"),
        ("6", "Trade di Q3 (menit 30-45)", "Reversal zone"),
    ]
    for vals in mistakes:
        row = write_row(ws, row, vals, s,
                        fonts=[s["red"], s["bold"], s["red"]],
                        fills=[s["red_fill"], None, None])

    return ws


def build_schedule_sheet(wb, s):
    ws = wb.create_sheet("Jadwal & Risk")
    ws.sheet_properties.tabColor = "E67E22"

    widths = [25, 25, 25, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    row = write_title(ws, row, "JADWAL WINDOW & RISK MANAGEMENT", 4, s)
    row += 1

    # JADWAL
    row = write_section(ws, row, "JADWAL WINDOW", 4, s)
    row = write_header(ws, row, ["Window", "Resolve (UTC)", "Resolve (WIB)", "Entry 1m before (WIB)"], s)

    sched = [
        ("15M", "XX:00, XX:15, XX:30, XX:45", "XX:07, XX:22, XX:37, XX:52", "XX:06, XX:21, XX:36, XX:51"),
        ("1H", "XX:00", "XX:07", "XX:06"),
    ]
    for vals in sched:
        row = write_row(ws, row, vals, s, fonts=[s["bold"], s["normal"], s["bold"], s["green"]])

    row += 1

    # CARA HITUNG GAP
    row = write_section(ws, row, "CARA HITUNG GAP", 4, s)
    row = write_header(ws, row, ["Step", "Action", "Contoh", ""], s)

    gap_steps = [
        ("1", "Lihat 'price to beat' di Polymarket UI", "$97,250", ""),
        ("2", "Buka Binance spot → harga BTC sekarang", "$97,380", ""),
        ("3", "Gap = |current - price_to_beat|", "$130", ""),
        ("4", "Cek tabel: gap $130 ≥ $75?", "YA → TRADE (bet UP)", ""),
    ]
    for vals in gap_steps:
        row = write_row(ws, row, vals, s, fonts=[s["bold"], s["normal"], s["bold"], s["normal"]])

    row += 1

    # POSITION SIZING
    row = write_section(ws, row, "POSITION SIZING (max 2-5% bankroll)", 4, s)
    row = write_header(ws, row, ["Bankroll", "Max per Trade", "Max Daily Loss", "Stop After"], s)

    sizing = [
        ("$200", "$10", "-$10 (5%)", "3 consecutive losses"),
        ("$500", "$25", "-$25 (5%)", "3 consecutive losses"),
        ("$1,000", "$50", "-$50 (5%)", "3 consecutive losses"),
        ("$2,000", "$100", "-$100 (5%)", "3 consecutive losses"),
    ]
    for vals in sizing:
        row = write_row(ws, row, vals, s, fonts=[s["bold"], s["green"], s["red"], s["orange"]])

    row += 1

    # LOSS LIMITS
    row = write_section(ws, row, "LOSS LIMITS", 4, s)
    row = write_header(ws, row, ["Timeframe", "Rule", "Action", ""], s)

    limits = [
        ("Daily", "3 consecutive losses", "STOP trading hari itu", ""),
        ("Daily", "-5% bankroll", "STOP trading hari itu", ""),
        ("Weekly", "-15% bankroll", "Review strategy", ""),
        ("Data Note", "Max loss streak = 4-6", "BISA terjadi!", ""),
    ]
    for vals in limits:
        row = write_row(ws, row, vals, s, fonts=[s["bold"], s["normal"], s["red"], s["normal"]],
                        fills=[None, None, s["red_fill"], None])

    row += 1

    # JANGAN PERNAH
    row = write_section(ws, row, "JANGAN PERNAH", 4, s)
    nevers = [
        ("All-in satu trade", "Satu loss = habis", "", ""),
        ("Martingale (double after loss)", "Exponential risk", "", ""),
        ("Revenge trade setelah loss streak", "Emotional = loss", "", ""),
        ("Pakai harga Hyperliquid tanpa buffer", "Deviasi $5-30", "", ""),
    ]
    for vals in nevers:
        row = write_row(ws, row, vals, s,
                        fonts=[s["red"], s["normal"], s["normal"], s["normal"]],
                        fills=[s["red_fill"], None, None, None])

    return ws


def main():
    wb = Workbook()
    s = make_styles()

    build_main_sheet(wb, s)
    build_rules_sheet(wb, s)
    build_schedule_sheet(wb, s)

    out = "/Users/administrator/Documents/Archantum/scripts/btc_cheatsheet.xlsx"
    wb.save(out)
    print(f"Saved to {out}")


if __name__ == "__main__":
    main()
